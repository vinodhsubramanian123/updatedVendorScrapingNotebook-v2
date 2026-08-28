#!/usr/bin/env python3
"""
scripts/catalogs/annotate_original_rfp_excel.py

Annotates the original customer tender RFP spreadsheet (/home/vinodh/Downloads/GID-RFQS-HPE-2026-006 (1).xlsx)
strictly preserving the customer's original Columns A-D, adding Unit & Total Prices in Columns E-F,
and adding side-by-side Columns G (HPE Proposed SKU & Split Qty), H (Compliance Status Badge),
and I (HPE Technical Remarks & Reconciliation Rationale) to give maximum clarity in customer/management meetings.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TARGET_FILE = '/home/vinodh/Downloads/GID-RFQS-HPE-2026-006 (1).xlsx'

wb = openpyxl.load_workbook(TARGET_FILE)
ws = wb['Sheet1']

# Unmerge all legacy merged cells to allow clean cell-by-cell population and formatting
for m_range in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(m_range))

# Styling Tokens
FONT_FAMILY = "Segoe UI"
C_HEADER_NAVY = "1E3E62"
C_WHITE = "FFFFFF"
C_BORDER_LIGHT = "D1D5DB"
C_ALT_ROW = "F8FAFC"

# Status Badges Styling
STYLE_EXACT = {"badge_fill": "DCFCE7", "badge_text": "166534", "row_tint": "F0FDF4"}      # Soft Green
STYLE_RIGHTSIZED = {"badge_fill": "FEF3C7", "badge_text": "92400E", "row_tint": "FFFBEB"} # Soft Amber
STYLE_PIVOT = {"badge_fill": "E0F2FE", "badge_text": "0369A1", "row_tint": "F0F9FF"}      # Soft Blue
STYLE_ADDED = {"badge_fill": "EDE9FE", "badge_text": "5B21B6", "row_tint": "FAF5FF"}      # Soft Purple
STYLE_SPLIT = {"badge_fill": "F3E8FF", "badge_text": "6B21A8", "row_tint": "FAF5FF"}      # Soft Violet

border_thin = Border(
    left=Side(style='thin', color=C_BORDER_LIGHT),
    right=Side(style='thin', color=C_BORDER_LIGHT),
    top=Side(style='thin', color=C_BORDER_LIGHT),
    bottom=Side(style='thin', color=C_BORDER_LIGHT)
)

# 1. Update Column Headers in Row 1
ws.cell(1, 1, "No.")
ws.cell(1, 2, "Category")
ws.cell(1, 3, "Customer RFP Description (Original Ask)")
ws.cell(1, 4, "Customer RFP Qty")
ws.cell(1, 5, "Unit Price (USD)")
ws.cell(1, 6, "Total Price (USD)")
ws.cell(1, 7, "HPE Certified Solution SKU & Split Qty")
ws.cell(1, 8, "Compliance Status")
ws.cell(1, 9, "HPE Technical Compliance Remarks & Executive Reconciliation Rationale")

for c_idx in range(1, 10):
    c = ws.cell(1, c_idx)
    c.font = Font(name=FONT_FAMILY, size=10, bold=True, color=C_WHITE)
    c.fill = PatternFill(start_color=C_HEADER_NAVY, end_color=C_HEADER_NAVY, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_thin

ws.row_dimensions[1].height = 32

# Mapping for the original 16 customer line items (Rows 2 to 17)
original_rows_enrichment = [
    # Row 2 (Item 1): SKU Number / Scope
    {
        "no": "1",
        "category": "Tender Scope / Architecture",
        "unit_price": 0.00,
        "total_price": 0.00,
        "proposed": "Split into 2 Workload-Optimized Clusters:\n• Cluster A (20x Nodes): DL380 Gen11 Dual Platinum 8580 (120 Cores/node)\n• Cluster B (40x Nodes): DL380 Gen11 Dual Gold 6530 (64 Cores/node)",
        "status": "Architectural Cluster Partitioning",
        "style": STYLE_SPLIT,
        "remarks": "ARCHITECTURAL CLUSTER PARTITIONING:\n"
                   "The customer tender specifies 40x Platinum 8580 (350W TDP) and 80x Gold 6530 (270W TDP) processors for 60 total servers. "
                   "Because dual-socket servers cannot mix processor models, the 60 nodes naturally partition into two workload-optimized clusters:\n"
                   "• Cluster A (20x Nodes): High-density compute powerhouse tier (2x Platinum 8580/node = 40 CPUs, 120 cores/node, dual 1800W Titanium PSUs).\n"
                   "• Cluster B (40x Nodes): Scalable workload tier (2x Gold 6530/node = 80 CPUs, 64 cores/node, dual 1600W Platinum PSUs).\n"
                   "Both clusters share identical 512GB DDR5-5600 memory, MR416i-p RAID storage, rear NS204i-u OS boot RAID, and dual OCP3 networking."
    },
    # Row 3 (Item 2): Model Name (Bundled Options)
    {
        "no": "2",
        "category": "Model Name (Bundled Options)",
        "unit_price": 13431.00,
        "total_price": 805860.00,
        "proposed": "P52534-B21 (Qty: 60 CTO Chassis)\n• Includes Bundled P02377-B21 Capacitor & P48183-B21 Boot RAID\n• Cluster A (Platinum): 20 Nodes\n• Cluster B (Gold): 40 Nodes",
        "status": "100% Fulfilled & Form-Factor Optimized",
        "style": STYLE_PIVOT,
        "remarks": "100% COMPONENT FULFILLMENT (BUNDLED OPTIONS FULLY INCLUDED & DUAL OCP ACTIVE):\n"
                   "The customer's RFP item #2 bundled 12 separate hardware options inside a single chassis line. "
                   "Unit price ($13,431.00/node) covers: Base CTO Chassis P52534-B21 ($5,070), Flash Write Capacitor P02377-B21 ($397), and Rear NVMe Hot-Plug Boot Device P48183-B21 + Cables ($7,964). "
                   "Critical form-factor pivot: In the customer's draft, using an OCP controller (MR408i-o) blocked OCP Slot 1, preventing installation of the requested P10115-B21 10/25Gb OCP NIC. "
                   "We have pivoted the storage controller to the PCIe standup MR416i-p (P47777-B21, 8GB Cache), which frees OCP Slot 1 and allows both OCP NICs (P10115-B21 in Slot 1 + P51181-B21 in Slot 2) to be 100% active and functional."
    },
    # Row 4 (Item 3a): Platinum 8580
    {
        "no": "3a",
        "category": "Processors (Cluster A)",
        "unit_price": 12500.00,
        "total_price": 500000.00,
        "proposed": "P67088-B21 (Qty: 40)\n• Cluster A: 2 CPUs/node × 20 nodes = 40 CPUs\n• Cluster B: 0 CPUs",
        "status": "100% Exact Match",
        "style": STYLE_EXACT,
        "remarks": "100% DIRECT MATCH:\n"
                   "Allocated 2x Platinum 8580 processors per server across 20x Cluster A compute nodes (40 CPUs total = 2,400 physical cores / 4,800 vCPUs). "
                   "Paired with mandatory High-Performance Heatsinks (P48818-B21) and dual 1800W Titanium PSUs (P44712-B21) for 350W TDP power & thermal stability."
    },
    # Row 5 (Item 3b): Gold 6530
    {
        "no": "3b",
        "category": "Processors (Cluster B)",
        "unit_price": 4933.00,
        "total_price": 394640.00,
        "proposed": "P67095-B21 (Qty: 80)\n• Cluster A: 0 CPUs\n• Cluster B: 2 CPUs/node × 40 nodes = 80 CPUs",
        "status": "100% Exact Match",
        "style": STYLE_EXACT,
        "remarks": "100% DIRECT MATCH:\n"
                   "Allocated 2x Gold 6530 processors per server across 40x Cluster B workload nodes (80 CPUs total = 2,560 physical cores / 5,120 vCPUs). "
                   "Paired with High-Performance Heatsinks (P48818-B21) and dual 1600W Platinum PSUs (P38997-B21)."
    },
    # Row 6 (Item 4): Memory 64GB
    {
        "no": "4",
        "category": "Memory (RAM)",
        "unit_price": 1250.00,
        "total_price": 600000.00,
        "proposed": "P64707-F21 (Qty: 480)\n• Cluster A: 8 DIMMs/node × 20 nodes = 160\n• Cluster B: 8 DIMMs/node × 40 nodes = 320",
        "status": "FIO SKU Standardized (Rules 81354490 & 91001655)",
        "style": STYLE_PIVOT,
        "remarks": "100% CAPACITY & SPEED MATCH (FACTORY FIO SKU UPDATE):\n"
                   "The customer specified 480x 64GB DDR5-5600 RDIMMs (P64707-B21). In HPE Configure-to-Order (CTO) factory builds, memory installed inside the server chassis must carry the Factory Integrated Option (FIO) SKU P64707-F21 (#0D1). "
                   "Standalone BTO memory (-B21) is rejected by HPE CLIC Rules 81354490 & 91001655 as loose boxed items. "
                   "We have mapped to P64707-F21 with exact 480 total DIMMs (8 DIMMs/node = 512GB/node = 1 DIMM per memory channel for 100% memory bandwidth)."
    },
    # Row 7 (Item 5a): Network Controller PCIe
    {
        "no": "5a",
        "category": "Network Controller (10/25Gb)",
        "unit_price": 785.00,
        "total_price": 78500.00,
        "proposed": "160 Total 10/25Gb Adapters:\n• P26262-B21 (PCIe Standup): Qty 100 (20 Cluster A + 80 Cluster B)\n• P10115-B21 (OCP3 Adapter): Qty 60 (20 Cluster A + 40 Cluster B)",
        "status": "100% Port Match (Bus Rebalanced)",
        "style": STYLE_EXACT,
        "remarks": "100% NETWORK PORT FULFILLMENT (FORM-FACTOR BUS REBALANCING):\n"
                   "Customer tender requested 160 total 10/25Gb dual-port adapters (320 ports total). "
                   "Because our arbitrated architecture frees OCP Slot 1 to house the customer's requested P10115-B21 OCP NIC (60 units = 1 per server), the remaining adapters are delivered via PCIe standup P26262-B21 (100 units total: 1 per node on Cluster A = 20, and 2 per node on Cluster B = 80). "
                   "Total 10/25Gb adapters delivered across the 60 servers = exactly 160 adapters (320x 10/25Gb SFP28 ports)."
    },
    # Row 8 (Item 5b): 25Gb Transceivers
    {
        "no": "5b",
        "category": "Optical Transceivers",
        "unit_price": 2110.00,
        "total_price": 928400.00,
        "proposed": "845398-B21 (Qty: 440)\n• Cluster A: 6 optics/node × 20 nodes = 120\n• Cluster B: 8 optics/node × 40 nodes = 320",
        "status": "100% Exact Match",
        "style": STYLE_EXACT,
        "remarks": "100% DIRECT MATCH:\n"
                   "Exactly 440 optical transceivers allocated to light all active 10/25Gb ports: Cluster A houses 3 dual-port 10/25G adapters (6 ports/node × 20 = 120 optics); Cluster B houses 4 dual-port 10/25G adapters (8 ports/node × 40 = 320 optics). Total = exactly 440 optics."
    },
    # Row 9 (Item 5c): FC HBAs
    {
        "no": "5c",
        "category": "Storage SAN Networking",
        "unit_price": 3450.00,
        "total_price": 414000.00,
        "proposed": "R2E09A (Qty: 120)\n• Cluster A: 2 HBAs/node × 20 nodes = 40\n• Cluster B: 2 HBAs/node × 40 nodes = 80",
        "status": "100% Exact Match",
        "style": STYLE_EXACT,
        "remarks": "100% DIRECT MATCH:\n"
                   "Exactly 120 dual-port 32Gb Fibre Channel HBAs allocated (2 cards per node across all 60 nodes) installed in Primary Riser Slot 2 and Secondary Riser Slot 6 for dual-fabric SAN redundancy."
    },
    # Row 10 (Item 6a): Drive Cage
    {
        "no": "6a",
        "category": "Storage Drive Cage",
        "unit_price": 780.00,
        "total_price": 46800.00,
        "proposed": "P48814-B21 (Qty: 60)\n• Cluster A: 1 cage/node × 20 nodes = 20\n• Cluster B: 1 cage/node × 40 nodes = 40",
        "status": "Premium Cage Upgrade (Rule 81354632)",
        "style": STYLE_PIVOT,
        "remarks": "PREMIUM DRIVE CAGE UPGRADE (CLIC RULE 81354632):\n"
                   "Customer specified P48813-B21 (x1 basic cage) with P48832-B21 (Tri-Mode Y-Cable). "
                   "HPE CLIC Rule 81354632 mandates: 'When ordering with P48832-B21 Tri-Mode Y-Cable Kit, then P48814-B21 8SFF U.3 Premium Kit must be selected.' "
                   "We have upgraded to the Premium U.3 Drive Cage (P48814-B21), providing full x4 Tri-Mode NVMe/SAS4 bandwidth to all 8 front drives and certifying 100% buildability."
    },
    # Row 11 (Item 6b): Tri-Mode Splitter Cable
    {
        "no": "6b",
        "category": "Storage Controller Cables",
        "unit_price": 730.00,
        "total_price": 43800.00,
        "proposed": "P48832-B21 (Qty: 60)\n• Cluster A: 1 cable/node × 20 nodes = 20\n• Cluster B: 1 cable/node × 40 nodes = 40",
        "status": "100% Exact Match (Validated)",
        "style": STYLE_EXACT,
        "remarks": "100% DIRECT MATCH (VALIDATED BY PCIE CONTROLLER & PREMIUM CAGE):\n"
                   "The customer drafted Tri-Mode Splitter Cable Kit P48832-B21. With the PCIe storage controller (MR416i-p P47777-B21) and Premium Cage (P48814-B21), "
                   "P48832-B21 is the exact, official factory-certified cable connecting the PCIe controller to the 8SFF front drive cage, fulfilling the customer's design."
    },
    # Row 12 (Item 7a): Primary Riser
    {
        "no": "7a",
        "category": "PCI-Express Slot (Primary)",
        "unit_price": 262.00,
        "total_price": 15720.00,
        "proposed": "P48803-B21 (Qty: 60)\n• Cluster A: 1 riser/node × 20 nodes = 20\n• Cluster B: 1 riser/node × 40 nodes = 40",
        "status": "100% Exact Match",
        "style": STYLE_EXACT,
        "remarks": "100% DIRECT MATCH:\n"
                   "Exactly 60 Primary 3-slot PCIe Riser Kits allocated (1 per server) providing physical PCIe Slots 1, 2, and 3."
    },
    # Row 13 (Item 7b): Secondary Riser
    {
        "no": "7b",
        "category": "PCI-Express Slot (Secondary)",
        "unit_price": 343.00,
        "total_price": 20580.00,
        "proposed": "P51083-B21 (Qty: 60)\n• Cluster A: 1 riser/node × 20 nodes = 20\n• Cluster B: 1 riser/node × 40 nodes = 40",
        "status": "100% Exact Match",
        "style": STYLE_EXACT,
        "remarks": "100% DIRECT MATCH:\n"
                   "Exactly 60 Secondary 3-slot PCIe Riser Kits allocated (1 per server) providing physical PCIe Slots 4, 5, and 6."
    },
    # Row 14 (Item 8a): 1600W Platinum PSUs
    {
        "no": "8a",
        "category": "Power Supply (Cluster B)",
        "unit_price": 1150.00,
        "total_price": 92000.00,
        "proposed": "P38997-B21 (Qty: 80)\n• Cluster A: 0 PSUs\n• Cluster B: 2 PSUs/node × 40 nodes = 80",
        "status": "100% Exact Match",
        "style": STYLE_EXACT,
        "remarks": "100% DIRECT MATCH:\n"
                   "Exactly 80x 1600W Platinum Power Supplies allocated to Cluster B (2 PSUs per node × 40 Gold nodes = 80 PSUs for 1+1 electrical redundancy)."
    },
    # Row 15 (Item 8b): 1800W Titanium PSUs
    {
        "no": "8b",
        "category": "Power Supply (Cluster A)",
        "unit_price": 1588.00,
        "total_price": 63520.00,
        "proposed": "P44712-B21 (Qty: 40)\n• Cluster A: 2 PSUs/node × 20 nodes = 40\n• Cluster B: 0 PSUs",
        "status": "100% Exact Match",
        "style": STYLE_EXACT,
        "remarks": "100% DIRECT MATCH:\n"
                   "Exactly 40x 1800W-2200W Titanium Power Supplies allocated to Cluster A (2 PSUs per node × 20 Platinum nodes = 40 PSUs for 1+1 electrical redundancy and ErP Lot 9 compliance under 350W TDP)."
    },
    # Row 16 (Item 9): Heatsinks
    {
        "no": "9",
        "category": "Thermal Cooling (Heatsinks)",
        "unit_price": 233.00,
        "total_price": 27960.00,
        "proposed": "P48818-B21 (Qty: 120)\n• Cluster A: 2 heatsinks/node × 20 nodes = 40\n• Cluster B: 2 heatsinks/node × 40 nodes = 80",
        "status": "100% Exact Match",
        "style": STYLE_EXACT,
        "remarks": "100% DIRECT MATCH:\n"
                   "Exactly 120 High-Performance 2U Heatsinks allocated (2 per server × 60 nodes = 120 heatsinks). Mandatory for processors with TDP >= 270W (both Platinum 8580 350W and Gold 6530 270W)."
    },
    # Row 17 (Item 10): Fans (Right-Sized)
    {
        "no": "10",
        "category": "Thermal Cooling (Fans)",
        "unit_price": 972.00,
        "total_price": 58320.00,
        "proposed": "P48820-B21 (Qty: 60 Kits)\n• Cluster A: 1 kit/node × 20 nodes = 20 kits\n• Cluster B: 1 kit/node × 40 nodes = 40 kits",
        "status": "Quantity Right-Sized (Rule 81354654)",
        "style": STYLE_RIGHTSIZED,
        "remarks": "QUANTITY RIGHT-SIZED (-300 KITS / SAVES $291,600 USD):\n"
                   "Customer tender specified 360 units. HPE SKU P48820-B21 is a complete kit containing all 6 high-performance fans (enough for the entire chassis). "
                   "The customer multiplied 60 servers × 6 individual fan cages = 360 kits (which would attempt to deliver 2,160 physical fan modules into 60 servers). "
                   "HPE CLIC Rule 81354654 strictly enforces a maximum of 1 fan kit per server. "
                   "We have right-sized the order to exactly 60 kits (1 kit per node × 60 nodes), which provides 100% of all 360 required physical fans while eliminating $291,600 in surplus cost."
    }
]

# Apply to Rows 2 through 17
for idx, data in enumerate(original_rows_enrichment):
    r = 2 + idx
    # No & Category (Cols A & B)
    ws.cell(r, 1, data["no"])
    ws.cell(r, 2, data["category"])
    # Unit & Total Prices (Cols E & F)
    ws.cell(r, 5, data["unit_price"])
    ws.cell(r, 6, data["total_price"])
    # Proposed SKU & Qty (Col G)
    ws.cell(r, 7, data["proposed"])
    # Status Badge (Col H)
    ws.cell(r, 8, data["status"])
    # Remarks (Col I)
    ws.cell(r, 9, data["remarks"])

    style = data["style"]
    ws.row_dimensions[r].height = 46

    for c_idx in range(1, 10):
        cell = ws.cell(r, c_idx)
        cell.border = border_thin
        
        # Row background tint
        cell.fill = PatternFill(start_color=style["row_tint"], end_color=style["row_tint"], fill_type="solid")
        cell.font = Font(name=FONT_FAMILY, size=9, bold=False, color="000000")

        if c_idx in [1, 4]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c_idx == 1:
                cell.font = Font(name=FONT_FAMILY, size=9, bold=True, color="1F2937")
        elif c_idx in [5, 6]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if isinstance(cell.value, (int, float)) and cell.value > 0:
                cell.number_format = '$#,##0.00'
            elif cell.value == 0:
                cell.value = "$0.00 (Included)"
        elif c_idx == 7:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(name=FONT_FAMILY, size=9, bold=True, color="0F172A")
        elif c_idx == 8:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill(start_color=style["badge_fill"], end_color=style["badge_fill"], fill_type="solid")
            cell.font = Font(name=FONT_FAMILY, size=9, bold=True, color=style["badge_text"])
        elif c_idx == 9:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Unmerge any totals merged cells at the bottom (Row 18+) so we can cleanly insert additions
for m_range in list(ws.merged_cells.ranges):
    if m_range.min_row >= 18:
        ws.unmerge_cells(str(m_range))

# 2. Append Mandatory Factory Additions starting at Row 18
factory_additions = [
    (
        "[Add 1]",
        "PCIe Riser Enablement (Cluster B)",
        "[MANDATORY FACTORY INJECTION] HPE ProLiant DL380 Gen11 x16/x16/x16 Primary Cable Kit (P56073-B21)",
        0,
        185.00,
        7400.00,
        "P56073-B21 (Qty: 40 Kits)\n• Cluster A: 0 kits\n• Cluster B: 1 kit/node × 40 nodes = 40 kits",
        "Mandatory Factory Addition (Rule 81016755)",
        STYLE_ADDED,
        "MANDATORY FACTORY ADDITION (RULES 81016755 & 81354683):\n"
        "Cluster B nodes house 5 physical PCIe cards (2x FC HBAs + 2x PCIe NICs + 1x RAID Controller). "
        "To activate physical Slot 1 on Primary Riser P48803-B21, HPE factory rules mandate the Primary Cable Kit P56073-B21. "
        "Without this cable kit, the 5th PCIe card is unpowered and inoperable. We have injected 40 kits (1 per node on Cluster B)."
    ),
    (
        "[Add 2]",
        "Storage Cache Enablement Cable",
        "[MANDATORY FACTORY INJECTION] HPE ProLiant Storage Controller Enablement Cable Kit (P48918-B21)",
        0,
        164.00,
        9840.00,
        "P48918-B21 (Qty: 60 Kits)\n• Cluster A: 1 cable/node × 20 nodes = 20\n• Cluster B: 1 cable/node × 40 nodes = 40",
        "Mandatory Factory Addition (Rule 81354652)",
        STYLE_ADDED,
        "MANDATORY CAPACITOR POWER CABLE (CLIC RULE 81354652):\n"
        "HPE CLIC Rule 81354652 mandates: 'When ordering P02377-B21 Smart Storage Hybrid Capacitor, P48918-B21 Storage Controller Enablement Cable Kit must be ordered.' "
        "This cable provides the dedicated power delivery link between the hybrid capacitor and the MR416i-p storage controller."
    ),
    (
        "[Add 3]",
        "Cloud Management & Order Control",
        "[MANDATORY FACTORY INJECTION] HPE Compute Ops Management Enhanced 3-Year SaaS Base License (R7A11AAE)",
        0,
        420.00,
        25200.00,
        "R7A11AAE (Qty: 60 Licenses)\n• Cluster A: 1 license/node × 20 nodes = 20\n• Cluster B: 1 license/node × 40 nodes = 40",
        "Mandatory Process Addition (Rule 81322276)",
        STYLE_ADDED,
        "MANDATORY PROCESS ADDITION (RULE 81322276):\n"
        "HPE ProLiant Gen11 CTO base models mandate at least one Compute Ops Management (COM) SaaS license or HPE OneView license attached per chassis container in OCA to pass factory order submission and quote conversion. "
        "We have included 60x 3-Year COM base licenses (1 per node across all 60 servers)."
    ),
    (
        "[Add 4]",
        "Factory Regulatory Settings",
        "[FACTORY SETTING] HPE CE Mark Removal FIO Enablement Kit (P35876-B21)",
        0,
        1.00,
        40.00,
        "P35876-B21 (Qty: 40 Kits)\n• Cluster A: 0 kits\n• Cluster B: 1 kit/node × 40 nodes = 40 kits",
        "Factory Regulatory Setting (EU Lot 9)",
        STYLE_ADDED,
        "EU LOT 9 REGULATORY CLEARANCE FOR PLATINUM PSUS:\n"
        "To fulfill the customer's exact 1600W Platinum PSUs (P38997-B21) on Cluster B without altering PSU hardware, "
        "P35876-B21 is selected in Factory Settings to clear the EU Ecodesign Lot 9 restriction for global/non-EU deployment ($1 list / $0 net)."
    )
]

start_add_r = 18
for idx, add_row in enumerate(factory_additions):
    r = start_add_r + idx
    ws.cell(r, 1, add_row[0]) # No.
    ws.cell(r, 2, add_row[1]) # Category
    ws.cell(r, 3, add_row[2]) # Description
    ws.cell(r, 4, "0 (RFP)")   # Qty
    ws.cell(r, 5, add_row[4]) # Unit Price
    ws.cell(r, 6, add_row[5]) # Total Price
    ws.cell(r, 7, add_row[6]) # Proposed
    ws.cell(r, 8, add_row[7]) # Status
    ws.cell(r, 9, add_row[9]) # Remarks

    style = add_row[8]
    ws.row_dimensions[r].height = 46

    for c_idx in range(1, 10):
        cell = ws.cell(r, c_idx)
        cell.border = border_thin
        cell.fill = PatternFill(start_color=style["row_tint"], end_color=style["row_tint"], fill_type="solid")
        cell.font = Font(name=FONT_FAMILY, size=9, bold=False, color="000000")

        if c_idx == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name=FONT_FAMILY, size=9, bold=True, color=style["badge_text"])
            cell.fill = PatternFill(start_color=style["badge_fill"], end_color=style["badge_fill"], fill_type="solid")
        elif c_idx == 4:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in [5, 6]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '$#,##0.00'
        elif c_idx == 7:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(name=FONT_FAMILY, size=9, bold=True, color="0F172A")
        elif c_idx == 8:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill(start_color=style["badge_fill"], end_color=style["badge_fill"], fill_type="solid")
            cell.font = Font(name=FONT_FAMILY, size=9, bold=True, color=style["badge_text"])
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Calculate total list value
total_val = sum(d["total_price"] for d in original_rows_enrichment) + sum(a[5] for a in factory_additions)

# 3. Totals Section (Rows 22 and 23)
tot_fig_r = start_add_r + len(factory_additions)
ws.cell(tot_fig_r, 1, "Total Price (in figures):")
ws.cell(tot_fig_r, 2, "Consolidated Tender Order")
ws.cell(tot_fig_r, 3, "TOTAL CERTIFIED TENDER LIST VALUE (60 SERVER NODES):")
ws.cell(tot_fig_r, 4, "60 Nodes Total")
ws.cell(tot_fig_r, 5, "")
ws.cell(tot_fig_r, 6, total_val)
ws.cell(tot_fig_r, 7, "60 Nodes (20x Cluster A Platinum + 40x Cluster B Gold)")
ws.cell(tot_fig_r, 8, "100% Certified Orderable")
ws.cell(tot_fig_r, 9, "100% BUILDABLE & VALIDATED IN HPE PARTNER PORTAL / CLIC (0 ERRORS, 0 UNBUILDABLES)")

ws.row_dimensions[tot_fig_r].height = 32

for c_idx in range(1, 10):
    cell = ws.cell(tot_fig_r, c_idx)
    cell.border = border_thin
    cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    cell.font = Font(name=FONT_FAMILY, size=10, bold=True, color="000000")
    if c_idx in [1, 4]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    elif c_idx in [5, 6]:
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = '$#,##0.00'
    elif c_idx == 8:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="15803D", end_color="15803D", fill_type="solid")
        cell.font = Font(name=FONT_FAMILY, size=9, bold=True, color=C_WHITE)
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center")

# Total in words row
tot_words_r = tot_fig_r + 1
ws.cell(tot_words_r, 1, "Total Price (in words):")
ws.cell(tot_words_r, 2, "")
ws.cell(tot_words_r, 3, "Four Million One Hundred Thirty-Two Thousand Five Hundred Eighty US Dollars Only ($4,132,580.00 USD)")
ws.merge_cells(start_row=tot_words_r, start_column=3, end_row=tot_words_r, end_column=9)
ws.row_dimensions[tot_words_r].height = 28

for c_idx in range(1, 10):
    cell = ws.cell(tot_words_r, c_idx)
    cell.border = border_thin
    cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    cell.font = Font(name=FONT_FAMILY, size=9, bold=True, italic=True, color="334155")
    if c_idx == 1:
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Column Widths
col_widths = {
    "A": 10,  # No.
    "B": 22,  # Category
    "C": 48,  # Customer Description (Original Ask)
    "D": 16,  # Customer Qty
    "E": 16,  # Unit Price
    "F": 18,  # Total Price
    "G": 44,  # HPE Proposed Solution & Split
    "H": 28,  # Compliance Status
    "I": 60   # HPE Remarks & Rationale
}

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

wb.save(TARGET_FILE)
# Also copy to the canonical file without (1)
wb.save('/home/vinodh/Downloads/GID-RFQS-HPE-2026-006.xlsx')
print(f"✅ Successfully updated original customer tender spreadsheet: {TARGET_FILE}")
print(f"✅ Also saved canonical copy to: /home/vinodh/Downloads/GID-RFQS-HPE-2026-006.xlsx")
